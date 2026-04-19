"""
Management command to import a MeiGray NHL population report Excel file.

Usage:
    python manage.py import_population_report <path_to_xlsx>

Reads 'POP REPORT BY TAG NUMBER' for jersey records and
'SET DATES - JERSEY SCHEDULES' for game-by-game schedule data.
Creates or updates MeiGrayEntry records keyed by tag number.
"""

import os
import re
import openpyxl
from django.core.files import File
from django.core.management.base import BaseCommand, CommandError

_END_SET_RE = re.compile(r'^End \w+ Set \S+$', re.IGNORECASE)

from memorabilia.models import MeiGrayEntry, PopulationReport


# Maps schedule jersey column values (may be uppercase) to the title-case
# labels used in 'End <Label> Set N' comments in the same sheet.
JERSEY_COMMENT_LABEL = {
    'White': 'White',
    'WHITE': 'White',
    'Blue': 'Blue',
    'BLUE': 'Blue',
    'THIRD': 'Third',
    'Third': 'Third',
    'Black': 'Black',
    'BLACK': 'Black',
    'Green': 'Green',
    'GREEN': 'Green',
    'Orange': 'Orange',
    'ORANGE': 'Orange',
    'Gold': 'Gold',
    'GOLD': 'Gold',
    'Red': 'Red',
    'RED': 'Red',
    'SS': 'SS',
}

# Maps substrings found in the pop report Color field to the schedule jersey values.
COLOR_TO_SCHEDULE = [
    ('third', 'Third'),
    ('black', 'Black'),
    ('white', 'White'),
    ('blue', 'Blue'),
    ('green', 'Green'),
    ('orange', 'Orange'),
    ('gold', 'Gold'),
    ('red', 'Red'),
    ('stadium series', 'SS'),
    ('ss', 'SS'),
]


def _detect_col_offset(rows, season):
    """
    Return the column index where data begins (0 or 1).

    Older files have a blank column A; newer files start at column A.
    We detect this by finding the season string or 'TAG #' in the first few rows.
    """
    for row in rows[:10]:
        for col_idx in range(min(3, len(row))):
            val = row[col_idx]
            if val and (season in str(val) or str(val).strip() == 'TAG #'):
                return col_idx
    return 0


def _parse_schedule(ws, season):
    """
    Parse the SET DATES sheet into a dict:
        (team_name_lower, schedule_jersey, set_number_int) -> [
            {"date": str, "opponent": str, "comment": str|None}, ...
        ]
    Team sections are identified by rows containing the season string.
    Column layout is auto-detected to support both old (offset=1) and new (offset=0) files.
    """
    rows = list(ws.iter_rows(values_only=True))
    schedule = {}

    col = _detect_col_offset(rows, season)

    # Find team section boundaries
    team_boundaries = []
    for i, row in enumerate(rows):
        if len(row) > col and row[col] and season in str(row[col]):
            team_boundaries.append(i)
    team_boundaries.append(len(rows))

    for idx, start in enumerate(team_boundaries[:-1]):
        end = team_boundaries[idx + 1]
        team_header = str(rows[start][col])
        team_name = team_header.replace(f' {season}', '').strip().title()

        color_set = {}      # jersey_value -> current set number (int)
        pending_advance = {}  # jersey_value -> True when next game starts new set

        for row in rows[start + 1:end]:
            if len(row) <= col + 2:
                continue
            date = row[col]
            opponent = row[col + 1]
            jersey = row[col + 2]
            comment = row[col + 3] if len(row) > col + 3 else None

            if not date or not opponent or not jersey:
                continue
            if str(date) in ('Date', 'PRESEASON', 'REGULAR SEASON', 'PLAYOFFS'):
                continue

            # Advance set counter for this jersey color if flagged
            if pending_advance.get(jersey):
                color_set[jersey] = color_set.get(jersey, 1) + 1
                pending_advance[jersey] = False

            current_set = color_set.get(jersey, 1)

            comment_str = str(comment) if comment else None
            if comment_str and _END_SET_RE.match(comment_str):
                comment_str = None

            key = (team_name.lower(), jersey, current_set)
            schedule.setdefault(key, []).append({
                'date': str(date),
                'opponent': str(opponent),
                'comment': comment_str,
            })

            # Flag next game of this jersey color as set N+1
            label = JERSEY_COMMENT_LABEL.get(jersey, jersey)
            if comment and f'End {label} Set' in str(comment):
                pending_advance[jersey] = True

    return schedule


def _color_to_schedule_jersey(color_str):
    """Map the pop report Color field to the schedule's jersey column value."""
    lower = color_str.lower()
    for substring, schedule_value in COLOR_TO_SCHEDULE:
        if substring in lower:
            return schedule_value
    return None


def _parse_set_number(raw):
    """Return an int set number, or None for promo/special sets."""
    try:
        return int(raw)
    except (TypeError, ValueError):
        return None


class Command(BaseCommand):
    help = 'Import MeiGray NHL population report from an Excel file'

    def add_arguments(self, parser):
        parser.add_argument('xlsx_path', type=str, help='Path to the .xlsx file')
        parser.add_argument('--season', type=str, required=True, help='Season label, e.g. 2024-25')

    def handle(self, *args, **options):
        path = options['xlsx_path']
        season = options['season']

        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except FileNotFoundError:
            raise CommandError(f'File not found: {path}')
        except Exception as e:
            raise CommandError(f'Could not open file: {e}')

        required = {'POP REPORT BY TAG NUMBER', 'SET DATES - JERSEY SCHEDULES'}
        missing = required - set(wb.sheetnames)
        if missing:
            raise CommandError(f'Missing sheets: {", ".join(missing)}')

        # Save the file to media storage and create/update the PopulationReport record
        self.stdout.write('Saving report file...')
        filename = os.path.basename(path)
        report = PopulationReport.objects.filter(season=season).first()
        with open(path, 'rb') as f:
            if report:
                report.file.delete(save=False)
                report.file.save(filename, File(f), save=True)
            else:
                report = PopulationReport(season=season)
                report.file.save(filename, File(f), save=True)

        self.stdout.write('Parsing schedule sheet...')
        schedule = _parse_schedule(wb['SET DATES - JERSEY SCHEDULES'], season)

        self.stdout.write('Importing pop report records...')
        ws = wb['POP REPORT BY TAG NUMBER']
        rows = list(ws.iter_rows(values_only=True))

        # Auto-detect column offset (older files have a blank column A).
        col = _detect_col_offset(rows, season)

        # Tag prefix varies by season (e.g. 'X' for 2024-25, 'W' for 2023-24, 'M' for 2015-16).
        data_rows = [r for r in rows[6:] if len(r) > col and r[col] and str(r[col])[:1].isalpha()]

        created = updated = 0

        for row in data_rows:
            fields = row[col:col + 8]
            tag, team, player, jsy_num, color, set_raw, size, notes = (list(fields) + [None] * 8)[:8]

            tag = str(tag).strip()
            team = str(team).strip() if team else ''
            player = str(player).strip() if player else ''
            color = str(color).strip() if color else ''
            size = str(size).strip() if size else ''
            notes = str(notes).strip() if notes else ''
            jsy_num = str(jsy_num).strip() if jsy_num is not None else ''
            set_num_int = _parse_set_number(set_raw)
            set_number_str = str(set_raw).strip() if set_raw is not None else ''

            # Look up games worn from the schedule
            games_worn = []
            sched_jersey = _color_to_schedule_jersey(color)
            if sched_jersey and set_num_int is not None:
                key = (team.lower(), sched_jersey, set_num_int)
                games_worn = schedule.get(key, [])

            defaults = {
                'team': team,
                'player': player,
                'jersey_number': jsy_num,
                'color': color,
                'set_number': set_number_str,
                'size': size,
                'notes': notes,
                'games_worn': games_worn,
                'report': report,
            }

            _, was_created = MeiGrayEntry.objects.update_or_create(
                tag_number=tag,
                defaults=defaults,
            )
            if was_created:
                created += 1
            else:
                updated += 1

        self.stdout.write(self.style.SUCCESS(
            f'Done. {created} created, {updated} updated from {len(data_rows)} records '
            f'(season: {season}).'
        ))
