"""
Shared logic for parsing and importing MeiGray population report Excel files.
Called by both the import_population_report management command and the admin action.
"""

import io
import re
from datetime import date, datetime

import openpyxl
from django.db import transaction

from memorabilia.models import MeiGrayEntry

_END_SET_RE = re.compile(r'^End \w+ Set \S+$', re.IGNORECASE)
_PRE_PLUS_RE = re.compile(r'(?:all\s+)?pre(?:season)?\s*\+\s*(.+)', re.IGNORECASE)

_TAG_HEADERS = {'TAG #', 'TAG NUMBER', 'TAG#'}

JERSEY_COMMENT_LABEL = {
    'White': 'White', 'WHITE': 'White',
    'Blue': 'Blue', 'BLUE': 'Blue',
    'THIRD': 'Third', 'Third': 'Third',
    'Black': 'Black', 'BLACK': 'Black',
    'Green': 'Green', 'GREEN': 'Green',
    'Orange': 'Orange', 'ORANGE': 'Orange',
    'Gold': 'Gold', 'GOLD': 'Gold',
    'Red': 'Red', 'RED': 'Red',
    'Yellow': 'Yellow', 'YELLOW': 'Yellow',
    'SS': 'SS',
}

COLOR_TO_SCHEDULE = [
    ('third', 'Third'),
    ('black', 'Black'),
    ('white', 'White'),
    ('blue', 'Blue'),
    ('green', 'Green'),
    ('orange', 'Orange'),
    ('gold', 'Gold'),
    ('red', 'Red'),
    ('yellow', 'Yellow'),
    ('stadium series', 'SS'),
    ('ss', 'SS'),
]

# Ordered by preference — first match wins.
_POP_SHEET_CANDIDATES = [
    'POP REPORT BY TAG NUMBER',
    'POP REPT BY TAG NUMBER',
    'POP REPORT SORTED BY INV TAG NU',
    'SORTED BY TAG NUMBER',
    'SORTED BY TAG #',
    'SORTED BY TAG#',
    'TAG NUMBER',
]

_UNUSED_POP_SHEETS = [
    'TEAM',
    'PLAYER',
    'SORTED BY TEAM',
    'SORTED BY PLAYER',
]

# Abbreviated team names in tag data → schedule words they should match.
_TEAM_NAME_ALIASES = {
    'vancanucks': ['vancouver'],
    'atlthrashers': ['atlanta'],
}


def _load_workbook(source):
    """Load an openpyxl workbook from a file path or bytes. Always uses read_only + data_only."""
    if isinstance(source, (str, bytes)):
        content = open(source, 'rb').read() if isinstance(source, str) else source
    else:
        content = source.read()
    return openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)


def _sheet_map(wb):
    """Case-insensitive mapping of sheet name → actual workbook name."""
    return {name.upper(): name for name in wb.sheetnames}


def _find_sheet(sheets, candidates):
    """Return the workbook sheet name for the first matching candidate, or None."""
    for candidate in candidates:
        if candidate in sheets:
            return sheets[candidate]
    return None


def _detect_col_offset(rows):
    for row in rows[:15]:
        for col_idx in range(min(3, len(row))):
            val = row[col_idx]
            if val and str(val).strip().upper() in _TAG_HEADERS:
                return col_idx
    return 0


def _find_header(rows, col):
    """Return (header_row_index, upper-cased header list)."""
    for i, row in enumerate(rows[:15]):
        val = row[col] if len(row) > col else None
        if val and str(val).strip().upper() in _TAG_HEADERS:
            return i, [str(v).strip().upper() if v else '' for v in row[col:col + 9]]
    return 5, []


def _short_season(season):
    """Convert 2024-2025 → 2024-25; pass 2024-25 through unchanged."""
    parts = season.split('-')
    if len(parts) == 2 and len(parts[1]) == 4:
        return f'{parts[0]}-{parts[1][2:]}'
    return season


def _is_modern_schedule(ws):
    """Return True if the sheet uses the modern per-game row format (has Date + Opponent headers)."""
    for row in ws.iter_rows(max_row=20, values_only=True):
        row_upper = {str(c).strip().upper() for c in row if c}
        if 'DATE' in row_upper and 'OPPONENT' in row_upper:
            return True
    return False


# ---------------------------------------------------------------------------
# Modern per-game schedule parser
# ---------------------------------------------------------------------------

def _parse_schedule(ws, season):
    rows = list(ws.iter_rows(values_only=True))
    schedule = {}
    short = _short_season(season)
    col = next(
        (col_idx for row in rows[:10] for col_idx in range(min(3, len(row)))
         if row[col_idx] and short in str(row[col_idx])),
        0
    )

    team_boundaries = []
    for i, row in enumerate(rows):
        if len(row) > col and row[col] and short in str(row[col]):
            team_boundaries.append(i)
    team_boundaries.append(len(rows))

    for idx, start in enumerate(team_boundaries[:-1]):
        end = team_boundaries[idx + 1]
        team_header = str(rows[start][col])
        team_name = team_header.replace(f' {short}', '').strip().title()

        color_set = {}
        pending_advance = {}

        for row in rows[start + 1:end]:
            if len(row) <= col + 2:
                continue
            date_val = row[col]
            opponent = row[col + 1]
            jersey = row[col + 2]
            comment = row[col + 3] if len(row) > col + 3 else None

            if not date_val or not opponent or not jersey:
                continue
            if str(date_val) in ('Date', 'PRESEASON', 'REGULAR SEASON', 'PLAYOFFS', 'BEST OF 3'):
                continue

            if pending_advance.get(jersey):
                color_set[jersey] = color_set.get(jersey, 1) + 1
                pending_advance[jersey] = False

            current_set = color_set.get(jersey, 1)

            comment_str = str(comment) if comment else None
            if comment_str and _END_SET_RE.match(comment_str):
                comment_str = None

            key = (team_name.lower(), jersey, current_set)
            schedule.setdefault(key, []).append({
                'date': str(date_val),
                'opponent': str(opponent),
                'comment': comment_str,
            })

            label = JERSEY_COMMENT_LABEL.get(jersey, jersey)
            if comment and f'End {label} Set' in str(comment):
                pending_advance[jersey] = True

    return schedule


# ---------------------------------------------------------------------------
# Team name resolution helpers (used by legacy parsers)
# ---------------------------------------------------------------------------

def _build_team_index(tag_teams):
    """Map normalized words → list of lowercased full team names for fuzzy resolution."""
    index = {}
    for team in tag_teams:
        team_lower = team.lower()
        for word in team_lower.split():
            if len(word) > 2:
                index.setdefault(word, []).append(team_lower)
        for extra_word in _TEAM_NAME_ALIASES.get(team_lower, []):
            index.setdefault(extra_word, []).append(team_lower)
    return index


def _resolve_teams(schedule_name, team_index):
    """
    Map a schedule team name to all matching tag team names (lowercased).
    Strips Home/Away suffix, then collects all tag teams that share a word with
    the schedule name. Returns a list — may contain multiple entries when the
    same city appears under different abbreviations (e.g. 'VanCanucks' and
    'Vancouver Canucks' both match 'Vancouver - Home').
    """
    name = re.sub(r'\s*-\s*(home|away|road).*', '', schedule_name, flags=re.IGNORECASE).strip()
    candidates = set()
    for word in name.lower().split():
        if len(word) > 2 and word in team_index:
            candidates.update(team_index[word])
    return list(candidates)


# ---------------------------------------------------------------------------
# Legacy third-jersey tab parser (individual game dates, no opponents)
# ---------------------------------------------------------------------------

def _parse_legacy_dates(cell_value, year_start):
    """
    Parse individual game dates from a cell like '11/9, 11/19, 12/14, 1/3'.
    Months >= 10 belong to year_start; months < 10 belong to year_start + 1.
    Returns a list of 'm/d/yyyy' date strings.
    """
    dates = []
    for m in re.finditer(r'\b(\d{1,2})/(\d{1,2})\b', cell_value):
        month = int(m.group(1))
        day = int(m.group(2))
        year = year_start if month >= 10 else year_start + 1
        dates.append(f'{month}/{day}/{year}')
    return dates


def _parse_schedule_legacy_thirds(rows, season, tag_teams):
    """
    Parse a legacy third jersey tab.
    Format: one row per team, 'Third Set 1' and 'Third Set 2' columns hold comma-separated dates.
    Returns {(team_lower, 'third', set_num): [game_dicts]}.
    """
    year_start = int(season.split('-')[0])
    team_index = _build_team_index(tag_teams)

    header_idx = third1_col = third2_col = None
    for i, row in enumerate(rows[:5]):
        for j, cell in enumerate(row):
            if not cell:
                continue
            cell_stripped = str(cell).strip().lower()
            if cell_stripped == 'third set 1':
                header_idx = i
                third1_col = j
            elif cell_stripped == 'third set 2':
                third2_col = j
        if header_idx is not None:
            break

    if header_idx is None or third1_col is None:
        return {}

    schedule = {}
    for row in rows[header_idx + 1:]:
        if not row[0]:
            continue
        raw_team = str(row[0]).strip()
        if not raw_team:
            continue

        full_teams = _resolve_teams(raw_team, team_index)
        if not full_teams:
            continue

        for set_num, col in [(1, third1_col), (2, third2_col)]:
            if col is None or len(row) <= col or not row[col]:
                continue
            dates = _parse_legacy_dates(str(row[col]), year_start)
            if dates:
                games = [{'date': d, 'opponent': None, 'comment': None} for d in dates]
                for full_team in full_teams:
                    schedule[(full_team, 'third', set_num)] = games

    return schedule


# ---------------------------------------------------------------------------
# Legacy home/road range tab parser (set date boundaries for NHL API lookup)
# ---------------------------------------------------------------------------

def _detect_set_columns(header_row):
    """
    Identify which column holds each set's date range.
    Returns {set_num: col_index} for sets 1-4.
    Skips game-count columns and metadata columns.
    """
    skip_keywords = ('games:', 'games in', 'promo', 'p/o set', 'total', 'playoffs')
    col_map = {}
    for j, cell in enumerate(header_row):
        if not cell:
            continue
        h = str(cell).lower()
        # Check the 3rd-set pattern before skip_keywords: its header legitimately
        # contains 'playoffs' (e.g. '3rd Regular Season / Playoffs:') which would
        # otherwise be caught by the skip gate.
        if '3rd' in h and ('regular' in h or '/ po' in h or 'playoff' in h):
            col_map[4] = j
        elif any(kw in h for kw in skip_keywords):
            continue
        elif 'preseason' in h:
            col_map[1] = j
        elif '1st' in h and 'regular' in h:
            col_map[2] = j
        elif '2nd' in h and 'regular' in h:
            col_map[3] = j
    return col_map


def _parse_range_value(cell_val):
    """
    Parse a spreadsheet date range cell into (start_date, end_date) or None.
    Handles: 'mm/dd/yy - mm/dd/yy', datetime objects, and 'All Preseason' strings.
    Returns:
      'preseason'              — plain preseason set
      ('preseason+', str)      — preseason + extra regular-season dates (2003-04 style)
      (start_date, end_date)   — explicit date range
      None                     — unparseable / skip
    """
    if cell_val is None:
        return None
    if isinstance(cell_val, (int, float)):
        return None  # game count cell

    if isinstance(cell_val, datetime):
        d = cell_val.date()
        return (d, d)

    s = str(cell_val).strip()
    if not s or s == '0':
        return None
    m = _PRE_PLUS_RE.match(s)
    if m:
        return ('preseason+', m.group(1).strip())
    if 'preseason' in s.lower() or 'all pre' in s.lower():
        return 'preseason'

    m = re.match(r'(\d{1,2})/(\d{1,2})/(\d{2,4})\s*-\s*(\d{1,2})/(\d{1,2})/(\d{2,4})', s)
    if m:
        def to_date(mo, d, y):
            y = int(y)
            if y < 100:
                y += 2000 if y < 50 else 1900
            return date(y, int(mo), int(d))
        return (to_date(m.group(1), m.group(2), m.group(3)),
                to_date(m.group(4), m.group(5), m.group(6)))
    return None


def _parse_range_rows(rows, team_index):
    """
    Parse data rows from a home/road range tab.
    Returns {(team_lower, home_away_str, set_num): range_value}.
    range_value is either 'preseason', (start_date, end_date), or a date range string.
    """
    # Find header row
    header_row = None
    header_idx = 0
    for i, row in enumerate(rows[:5]):
        if row[0] and 'team' in str(row[0]).lower():
            header_row = row
            header_idx = i
            break
    if header_row is None:
        return {}

    col_map = _detect_set_columns(header_row)
    if not col_map:
        return {}

    ranges = {}
    for row in rows[header_idx + 1:]:
        if not row[0]:
            continue
        raw_team = str(row[0]).strip()
        if not raw_team or raw_team.lower().startswith('team'):
            continue

        if re.search(r'-\s*home', raw_team, re.IGNORECASE):
            home_away = 'home'
        elif re.search(r'-\s*(?:away|road)', raw_team, re.IGNORECASE):
            home_away = 'away'
        else:
            continue

        full_teams = _resolve_teams(raw_team, team_index)
        if not full_teams:
            continue

        for set_num, col in col_map.items():
            if col >= len(row):
                continue
            parsed = _parse_range_value(row[col])
            if parsed is not None:
                for full_team in full_teams:
                    ranges[(full_team, home_away, set_num)] = parsed

    return ranges


def _build_schedule(wb, sheets, pop_sheet_name, tag_teams, season):
    """
    Scan every non-population-report sheet and build a combined schedule dict.
    Sheet type is detected by content, not by name, so new report formats are
    handled automatically.
    """
    team_index = _build_team_index(tag_teams)
    schedule = {}
    set_ranges = {}

    skip_uppers = {pop_sheet_name.upper()} | {s.upper() for s in _UNUSED_POP_SHEETS}

    for sheet_upper, sheet_name in sheets.items():
        if sheet_upper in skip_uppers:
            continue
        ws = wb[sheet_name]
        if _is_modern_schedule(ws):
            schedule.update(_parse_schedule(wb[sheet_name], season))
        else:
            rows = list(ws.iter_rows(values_only=True))
            ranges = _parse_range_rows(rows, team_index)
            if ranges:
                set_ranges.update(ranges)
            thirds = _parse_schedule_legacy_thirds(rows, season, tag_teams)
            schedule.update(thirds)

    if set_ranges:
        schedule.update(_build_schedule_from_ranges_only(set_ranges, season))

    return schedule


# ---------------------------------------------------------------------------

def _jersey_home_away(color_str):
    """
    Return 'third', 'away', or 'home' based on the jersey color string.
    White jerseys are away; Third jerseys are thirds; everything else is home.
    """
    lower = color_str.lower()
    if 'third' in lower:
        return 'third'
    if 'white' in lower:
        return 'away'
    return 'home'


def _build_schedule_from_ranges_only(set_ranges, season=None):
    """
    Build a schedule from range data alone (no API).
    Returns {(team_lower, 'home'|'away', set_num): [entry]} where each entry is
    one of three typed dicts:
      {'type': 'range',     'start': 'YYYY-MM-DD', 'end': 'YYYY-MM-DD'}
      {'type': 'preseason'}
      {'type': 'preseason+', 'extra_dates': ['YYYY-MM-DD', ...]}
    """
    schedule = {}
    for (team_lower, home_away, set_num), range_val in set_ranges.items():
        if range_val == 'preseason':
            games = [{'type': 'preseason'}]
        elif isinstance(range_val, tuple) and range_val[0] == 'preseason+':
            year_start = int(season.split('-')[0]) if season else 2000
            extra_strs = _parse_legacy_dates(range_val[1], year_start)
            extra_dates = []
            for ds in extra_strs:
                mo, d, y = ds.split('/')
                extra_dates.append(f'{y}-{int(mo):02d}-{int(d):02d}')
            games = [{'type': 'preseason+', 'extra_dates': extra_dates}]
        else:
            start_d, end_d = range_val
            games = [{'type': 'range', 'start': str(start_d), 'end': str(end_d)}]
        schedule[(team_lower, home_away, set_num)] = games
    return schedule


# ---------------------------------------------------------------------------
# Color/jersey lookup (used by both tag parsers)
# ---------------------------------------------------------------------------

def _color_to_schedule_jersey(color_str):
    lower = color_str.lower()
    for substring, schedule_value in COLOR_TO_SCHEDULE:
        if substring in lower:
            return schedule_value
    return None


def _lookup_games(schedule, team, color, set_num_int, set_raw=None):
    """
    Look up games_worn from schedule dict.
    Tries modern color-label key first, then home/away key (API-derived).
    """
    if set_num_int is None:
        return []

    # Modern format: key uses jersey color label (White, Blue, Third, etc.)
    sched_jersey = _color_to_schedule_jersey(color)
    if sched_jersey:
        games = schedule.get((team.lower(), sched_jersey, set_num_int))
        if games is not None:
            return games

    # API/range format: key uses home/away/third.
    # In legacy range schedules (2002-03, 2003-04), preseason occupies key 1 and
    # regular sets are keyed 2-4, but the tag data numbers regular sets 1-3 with
    # a "(Reg)" suffix. Add 1 to align. Third jerseys are keyed consistently and
    # need no offset.
    ha = _jersey_home_away(color)
    if ha != 'third' and set_raw and '(Reg)' in str(set_raw):
        range_key = set_num_int + 1
    else:
        range_key = set_num_int
    return schedule.get((team.lower(), ha, range_key), [])


# ---------------------------------------------------------------------------
# Set number parsing
# ---------------------------------------------------------------------------

def _parse_set_number(raw):
    """Return integer set number, handling formats like '2 (Reg)', '1 (Pre)', or plain 2."""
    if raw is None:
        return None
    try:
        return int(raw)
    except (TypeError, ValueError):
        m = re.search(r'\d+', str(raw))
        return int(m.group()) if m else None


# ---------------------------------------------------------------------------
# Tag data parsers
# ---------------------------------------------------------------------------

def _parse_entries_color_set(rows, col, header_idx, league, schedule, report):
    """
    8-column format: TAG #, Team, Player, JSY #, Color (or Jersey Type), Set, Size, Notes
    Used by NHL and similar reports with separate color and set number columns.
    Returns (entries, total_rows).
    """
    data_rows = [
        r for r in rows[header_idx + 1:]
        if len(r) > col and r[col] and str(r[col])[:1].isalpha()
    ]
    entries = []
    for row in data_rows:
        fields = (list(row[col:col + 8]) + [None] * 8)[:8]
        tag = str(fields[0]).strip() if fields[0] else ''
        team = str(fields[1]).strip() if fields[1] else ''
        player = str(fields[2]).strip() if fields[2] else ''
        jsy_num = str(fields[3]).strip() if fields[3] is not None else ''
        color = str(fields[4]).strip() if fields[4] else ''
        set_raw = fields[5]
        size = str(fields[6]).strip() if fields[6] else ''
        notes = str(fields[7]).strip() if fields[7] else ''

        set_num_int = _parse_set_number(set_raw)
        set_number_str = str(set_raw).strip() if set_raw is not None else ''

        games_worn = _lookup_games(schedule, team, color, set_num_int, set_raw=set_raw)

        entries.append(MeiGrayEntry(
            tag_number=tag,
            league=league,
            team=team,
            player=player,
            jersey_number=jsy_num,
            color=color,
            set_number=set_number_str,
            size=size,
            notes=notes,
            games_worn=games_worn,
            report=report,
        ))
    return entries, len(data_rows)


def _parse_entries_set_only(rows, col, header_idx, league, schedule, report):
    """
    6-column format: Tag Number, Team, Player Name, JSY #, Set (color description), Size
    Used by NWHL and similar reports where color and set are combined in one column.
    Returns (entries, total_rows).
    """
    data_rows = [
        r for r in rows[header_idx + 1:]
        if len(r) > col and r[col] and str(r[col])[:1].isalpha()
    ]
    entries = []
    for row in data_rows:
        fields = (list(row[col:col + 6]) + [None] * 6)[:6]
        tag = str(fields[0]).strip() if fields[0] else ''
        team = str(fields[1]).strip() if fields[1] else ''
        player = str(fields[2]).strip() if fields[2] else ''
        jsy_num = str(fields[3]).strip() if fields[3] is not None else ''
        color = str(fields[4]).strip() if fields[4] else ''
        size = str(fields[5]).strip() if fields[5] else ''

        # 6-column format has no set number; always treat as set 1
        games_worn = _lookup_games(schedule, team, color, 1)

        entries.append(MeiGrayEntry(
            tag_number=tag,
            league=league,
            team=team,
            player=player,
            jersey_number=jsy_num,
            color=color,
            set_number='',
            size=size,
            notes='',
            games_worn=games_worn,
            report=report,
        ))
    return entries, len(data_rows)


# ---------------------------------------------------------------------------
# Dry-run analysis (no database writes)
# ---------------------------------------------------------------------------

def analyze_file(path, season, league):
    """
    Parse an XLSX at *path* and return verbose diagnostics without touching the database.
    Returns a dict with:
      total_raw       — row count after the header row
      total_parsed    — rows that passed the tag filter
      unique          — unique entries after deduplication
      skipped         — [(row_num, raw_tag, reason), ...]
      duplicates      — [tag_number, ...]  (last occurrence was kept)
      with_dates      — entries whose games_worn is non-empty
      without_dates   — entries whose games_worn is empty
      no_dates_breakdown — Counter {(team, set_number): count}
    """
    from collections import Counter

    wb = _load_workbook(path)
    sheets = _sheet_map(wb)

    pop_sheet_name = _find_sheet(sheets, _POP_SHEET_CANDIDATES)
    if not pop_sheet_name:
        raise ValueError(
            f'No population report sheet found. Sheets present: {", ".join(wb.sheetnames)}'
        )

    ws = wb[pop_sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    col = _detect_col_offset(rows)
    header_idx, headers = _find_header(rows, col)

    tag_teams = {
        str(r[col + 1]).strip()
        for r in rows[header_idx + 1:]
        if len(r) > col + 1 and r[col + 1] and str(r[col + 1])[:1].isalpha()
    }
    print(f'Identified tag teams for schedule resolution: {tag_teams}\n')

    schedule = _build_schedule(wb, sheets, pop_sheet_name, tag_teams, season)

    # Identify skipped rows before the tag filter is applied
    all_rows = rows[header_idx + 1:]
    skipped = []
    for i, r in enumerate(all_rows):
        row_num = header_idx + 2 + i  # 1-indexed spreadsheet row number
        raw = r[col] if len(r) > col else None
        if not raw:
            skipped.append((row_num, raw, 'empty tag'))
        elif not str(raw)[:1].isalpha():
            skipped.append((row_num, raw, f'tag does not start with a letter: {str(raw)!r}'))

    # Parse entries — report=None is safe; FK is nullable and nothing is saved
    if 'COLOR' in headers or 'JERSEY TYPE' in headers:
        entries, _ = _parse_entries_color_set(rows, col, header_idx, league, schedule, None)
    else:
        entries, _ = _parse_entries_set_only(rows, col, header_idx, league, schedule, None)

    seen = {}
    duplicates = []
    for e in entries:
        if e.tag_number in seen:
            duplicates.append(e.tag_number)
        seen[e.tag_number] = e
    unique_entries = list(seen.values())

    with_dates = sum(1 for e in unique_entries if e.games_worn)
    without_dates = len(unique_entries) - with_dates
    no_dates_breakdown = Counter(
        (e.team, e.set_number) for e in unique_entries if not e.games_worn
    )

    return {
        'total_raw': len(all_rows),
        'total_parsed': len(entries),
        'unique': len(unique_entries),
        'skipped': skipped,
        'duplicates': duplicates,
        'with_dates': with_dates,
        'without_dates': without_dates,
        'no_dates_breakdown': no_dates_breakdown,
    }


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def import_entries(report):
    """
    Parse the XLSX stored on report.file and replace all MeiGrayEntry records for
    this report. Runs in a transaction so a parse failure leaves existing data intact.
    Returns (deleted_count, created_count, total_rows, duplicates).
    Raises ValueError if the population report sheet is missing.
    """
    season = report.season
    league = report.league

    with report.file.open('rb') as f:
        content = f.read()

    wb = _load_workbook(content)
    sheets = _sheet_map(wb)

    pop_sheet_name = _find_sheet(sheets, _POP_SHEET_CANDIDATES)
    if not pop_sheet_name:
        raise ValueError(
            f'No population report sheet found. Sheets present: {", ".join(wb.sheetnames)}'
        )

    ws = wb[pop_sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    col = _detect_col_offset(rows)
    header_idx, headers = _find_header(rows, col)

    # Extract team names from tag data for fuzzy resolution
    tag_teams = {
        str(r[col + 1]).strip()
        for r in rows[header_idx + 1:]
        if len(r) > col + 1 and r[col + 1] and str(r[col + 1])[:1].isalpha()
    }

    schedule = _build_schedule(wb, sheets, pop_sheet_name, tag_teams, season)

    # --- Parse tag entries ---
    if 'COLOR' in headers or 'JERSEY TYPE' in headers:
        entries, total_rows = _parse_entries_color_set(rows, col, header_idx, league, schedule, report)
    else:
        entries, total_rows = _parse_entries_set_only(rows, col, header_idx, league, schedule, report)

    seen = {}
    duplicates = []
    for e in entries:
        if e.tag_number in seen:
            duplicates.append(e.tag_number)
        seen[e.tag_number] = e
    unique_entries = list(seen.values())

    with_dates = sum(1 for e in unique_entries if e.games_worn)
    without_dates = len(unique_entries) - with_dates

    with transaction.atomic():
        deleted, _ = MeiGrayEntry.objects.filter(report=report).delete()
        MeiGrayEntry.objects.bulk_create(unique_entries)

    return deleted, len(unique_entries), total_rows, duplicates, with_dates, without_dates
