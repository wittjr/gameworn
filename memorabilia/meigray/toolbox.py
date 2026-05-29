"""
Shared, format-agnostic primitives for parsing MeiGray population reports.

Nothing in here is year-specific. Per-year parsers (parsers_YYYY.py) compose
these building blocks; the registry decides which parsers a given season uses.
"""

import io
import re
from datetime import date, datetime
import sys

import openpyxl

from memorabilia.models import MeiGrayTagEntry

_TAG_HEADERS = {'TAG #', 'TAG NUMBER', 'TAG#'}

# Canonical team name -> alternate names that may appear in tag/schedule sheets.
_TEAM_VARIANTS = {
    'New York Rangers':        ['NY Rangers'],
    'New York Islanders':      ['NY Islanders'],
    'New Jersey Devils':       ['NJ Devils'],
    'Vancouver Canucks':       ['Vancanucks', 'Vancouver'],
    'Atlanta Thrashers':       ['Atl Thrashers', 'Atl Thrash', 'AtlThrash', 'Atlanta'],
    'St. Louis Blues':         ['St Louis Blues', 'St. Louis', 'St Louis'],
    'Montreal Canadiens':      ['Montréal Canadiens', 'Montreal'],
    'Mighty Ducks of Anaheim': ['Anaheim'],
    'Boston Bruins':           ['Boston'],
    'Buffalo Sabres':          ['Buffalo'],
    'Calgary Flames':          ['Calgary'],
    'Carolina Hurricanes':     ['Carolina'],
    'Chicago Blackhawks':      ['Chicago'],
    'Colorado Avalanche':      ['Colorado'],
    'Columbus Blue Jackets':   ['Columbus'],
    'Dallas Stars':            ['Dallas'],
    'Detroit Red Wings':       ['Detroit'],
    'Edmonton Oilers':         ['Edmonton'],
    'Florida Panthers':        ['Florida'],
    'Los Angeles Kings':       ['LA Kings'],
    'Minnesota Wild':          ['Minnesota'],
    'Nashville Predators':     ['Nashville'],
    'Ottawa Senators':         ['Ottawa'],
    'Philadelphia Flyers':     ['Philadelphia'],
    'Phoenix Coyotes':         ['Phoenix'],
    'Pittsburgh Penguins':     ['Pittsburgh'],
    'San Jose Sharks':         ['San Jose'],
    'Tampa Bay Lightning':     ['Tampa Bay'],
    'Toronto Maple Leafs':     ['Toronto'],
    'Washington Capitals':     ['Washington'],
}

# Built once at import; lowercased variant -> canonical name.
_VARIANT_TO_CANONICAL = {k.lower(): k for k in _TEAM_VARIANTS}
_VARIANT_TO_CANONICAL.update(
    {v.lower(): k for k, vs in _TEAM_VARIANTS.items() for v in vs}
)


# ---------------------------------------------------------------------------
# Workbook loading + sheet-name normalization
# ---------------------------------------------------------------------------

def load_workbook(source):
    """Load an openpyxl workbook from a path or bytes. Always read_only + data_only."""
    if isinstance(source, (str, bytes)):
        content = open(source, 'rb').read() if isinstance(source, str) else source
    else:
        content = source.read()
    return openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)


def short_season(season):
    """Normalize '2002-2003' -> '2002-03'; pass '2002-03' through unchanged."""
    parts = str(season).split('-')
    if len(parts) == 2 and len(parts[1]) == 4:
        return f'{parts[0]}-{parts[1][2:]}'
    return season


def norm_sheet(name):
    """Normalize a sheet name for matching: upper-case, collapse internal
    whitespace, strip ends. Tolerates trailing-space / casing drift across years."""
    return re.sub(r'\s+', ' ', str(name)).strip().upper()


def sheet_lookup(wb):
    """{normalized name -> actual workbook sheet name}."""
    return {norm_sheet(n): n for n in wb.sheetnames}


# ---------------------------------------------------------------------------
# Tag-sheet header / column detection
# ---------------------------------------------------------------------------

def detect_col_offset(rows):
    for row in rows[:15]:
        for col_idx in range(min(3, len(row))):
            val = row[col_idx]
            if val and str(val).strip().upper() in _TAG_HEADERS:
                return col_idx
    return 0


def find_header(rows, col):
    """Return (header_row_index, upper-cased header list)."""
    for i, row in enumerate(rows[:15]):
        val = row[col] if len(row) > col else None
        if val and str(val).strip().upper() in _TAG_HEADERS:
            return i, [str(v).strip().upper() if v else '' for v in row[col:col + 12]]
    return 5, []


def format_set_number(raw):
    """Display string for a Set cell. A datetime/date is Excel's mis-coercion
    of a string like '3-Jan' -> restore it as '3-Jan' rather than a timestamp."""
    if raw is None:
        return ''
    if isinstance(raw, (datetime, date)):
        return f"{raw.day}-{raw.strftime('%b')}"
    return str(raw).strip()


# ---------------------------------------------------------------------------
# Tag-row -> MeiGrayTagEntry
# ---------------------------------------------------------------------------

def read_tag_rows(wb, actual, manifest, report):
    """Locate, slice and parse the tag sheet. Shared by every year's parser."""
    from memorabilia.meigray.registry import sheet
    rows = list(wb[sheet(actual, manifest.tag)].iter_rows(values_only=True))
    col = detect_col_offset(rows)
    header_idx, headers = find_header(rows, col)
    return parse_tag_sheet(rows, col, header_idx, headers, report)


def collect_tag_teams(tag_entries):
    """Set of team names seen on the tag sheet, for matching against schedule
    team headers."""
    return {e.team for e in tag_entries if e.team}


def read_schedule_sheet(wb, actual, sheet_name):
    """Load a schedule sheet's rows. Returns [] if the sheet is absent."""
    from memorabilia.meigray.registry import sheet
    return list(wb[sheet(actual, sheet_name)].iter_rows(values_only=True))


def read_combined_schedule(wb, actual, sheet_name, season, tag_teams, set_dates_prefix=None):
    """Default modern reader: try the per-team-block summary table for
    set-range rows AND the per-game block for game rows on the same sheet.
    Either may yield zero rows depending on the year's layout."""
    rows = read_schedule_sheet(wb, actual, sheet_name)
    games = parse_pergame_schedule(rows, tag_teams)
    set_ranges = []
    if set_dates_prefix is not None:
        set_ranges = parse_set_dates_table(rows, season, tag_teams, set_dates_prefix)
    return games, set_ranges


def parse_tag_sheet(rows, col, header_idx, headers, report):
    """Read the POP REPORT BY TAG NUMBER sheet into MeiGrayTagEntry rows.
    Layout: TAG #, Team, Player, JSY #, Color, Set, Size, [Version,] Comments.
    Tag..Size are at fixed offsets col+0..col+6; the comment column is located
    from the header (some years insert a 'Version' column before it)."""
    hdr = rows[header_idx] if header_idx < len(rows) else ()
    comment_col = None
    for j in range(col, min(len(hdr), col + 12)):
        v = hdr[j]
        if v and ('CUSTOMIZATION' in str(v).upper() or 'COMMENT' in str(v).upper()):
            comment_col = j
            break

    entries = []
    seen = set()
    for row in rows[header_idx + 1:]:
        if len(row) <= col or not row[col] or not str(row[col])[:1].isalpha():
            continue
        fields = (list(row[col:col + 7]) + [None] * 7)[:7]
        tag = str(fields[0]).strip() if fields[0] else ''
        if not tag or tag in seen:
            continue
        seen.add(tag)
        team = str(fields[1]).strip() if fields[1] else ''
        player = str(fields[2]).strip() if fields[2] else ''
        jsy_num = str(fields[3]).strip() if fields[3] is not None else ''
        color = str(fields[4]).strip() if fields[4] else ''
        set_number_str = format_set_number(fields[5])
        size = str(fields[6]).strip() if fields[6] else ''
        comment = ''
        if comment_col is not None and comment_col < len(row) and row[comment_col]:
            comment = str(row[comment_col]).strip()

        entries.append(MeiGrayTagEntry(
            tag_number=tag,
            season=report.season,
            league=report.league,
            team=team,
            player=player,
            jersey_number=jsy_num,
            color=color,
            set_number=set_number_str,
            size=size,
            notes=comment[:500],
            report=report,
        ))
    return entries


# ---------------------------------------------------------------------------
# Team-name resolution (fuzzy match schedule names <-> tag team names)
# ---------------------------------------------------------------------------

def build_team_index(tag_teams):
    """Map normalized name (canonical + every known variant) -> tag team names."""
    index = {}
    for team in tag_teams:
        canonical = _VARIANT_TO_CANONICAL.get(team.lower(), team)
        names = {canonical, *_TEAM_VARIANTS.get(canonical, []), team}
        for n in names:
            index.setdefault(n.lower(), []).append(team)
    return index


def resolve_teams(schedule_name, team_index):
    """Map a schedule team name to all matching tag team names.
    Strips a Home/Away/Road suffix, then collects every tag team sharing a word."""
    name = re.sub(r'\s*-\s*(home|away|road).*', '', schedule_name, flags=re.IGNORECASE).strip()
    candidates = set()
    if name.lower() in team_index:
        candidates.update(team_index[name.lower()])
    return list(candidates)


# ---------------------------------------------------------------------------
# Date parsing
# ---------------------------------------------------------------------------

_WD_RE = r'(?:mon|tue|tues|wed|thu|thur|thurs|fri|sat|sun)[a-z]*'
_YEAR_END_MONTH_RE = r'.* (?:sep|oct|nov|dec) .*'



def parse_schedule_date(val):
    """A schedule date cell -> 'YYYY-MM-DD', or None. Handles datetime cells,
    weekday-trailing ('Oct 5 2005, Wed') and weekday-leading ('Sat Jan 19, 2013')."""
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        return (val.date() if isinstance(val, datetime) else val).isoformat()
    s = str(val).strip()
    if not s or s.upper() == 'DATE':
        return None
    s = re.sub(r'^%s\.?,?\s+' % _WD_RE, '', s, flags=re.IGNORECASE)
    s = re.sub(r',?\s*%s\.?\s*$' % _WD_RE, '', s, flags=re.IGNORECASE)
    s = re.sub(r'([A-Za-z])\.', r'\1 ', s)
    s = re.sub(r'\s+,', ',', re.sub(r'\s+', ' ', s)).strip().rstrip(',').strip()
    s = s.replace('Sept', 'September')
    for fmt in ('%b %d %Y', '%b %d, %Y', '%B %d %Y', '%B %d, %Y',
                '%d %b %Y', '%d %b, %Y', '%d %B %Y', '%d %B, %Y'):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            pass
    return None


# ---------------------------------------------------------------------------
# Per-game schedule reader (modern format: one row per game)
# ---------------------------------------------------------------------------

_TEAM_HDR_RE = re.compile(
    r'^(?!preseason|regular\s+season).*\b20\d\d'
    r'(?:(?:-\d\d|-20\d\d)(?:\s+season)?|(?:-\d\d|-20\d\d)?\s+season)\s*$',
    re.IGNORECASE)
_HDR_SUFFIX_RE = re.compile(
    r'\s*\b20\d\d(?:-\d\d|-20\d\d)?(?:\s+season)?\b.*$', re.IGNORECASE)

_SECTION_LABELS = ('preseason', 'regularseason', 'playoffs', 'exhibition', 'round-robin')


def _hdr_role_cols(header_row, date_cols):
    """For each DATE column, map its group's role columns. Returns
    {date_col: (jersey_col, note_col, opp_col)}. Detects 'JSY'/'Jersey',
    'Opponent', 'NOTES' from the header; falls back to fixed offsets."""
    roles = {}
    for idx, dc in enumerate(date_cols):
        nxt = date_cols[idx + 1] if idx + 1 < len(date_cols) else len(header_row)
        jcol = ocol = ncol = None
        for c in range(dc + 1, nxt):
            h = str(header_row[c]).strip().upper() if c < len(header_row) \
                and header_row[c] is not None else ''
            if h in ('JSY', 'JERSEY') and jcol is None:
                jcol = c
            elif h in ('OPPONENT', 'OPP') and ocol is None:
                ocol = c
            elif h in ('NOTES', 'NOTE', 'COMMENTS', 'COMMENT') and ncol is None:
                ncol = c
        if jcol is None:
            jcol, ncol, ocol = dc + 1, dc + 2, dc + 3
        roles[dc] = (jcol, ncol if ncol is not None else dc + 2,
                     ocol if ocol is not None else dc + 3)
    return roles


def _jersey_is_home(jersey):
    """For the team whose block we're in: a White/Away jersey means we're on
    the road; everything else means home."""
    return 'white' not in jersey.lower() and 'away' not in jersey.lower()


def parse_pergame_schedule(rows, tag_teams):
    """Parse the per-game schedule sheet into a flat list of game dicts:
        {'team', 'date', 'opponent', 'jersey', 'home_game', 'comment'}
    Each per-team block starts with a team-header row carrying the season suffix.
    Drops rows that look like set-end markers or section labels."""
    team_index = build_team_index(tag_teams)

    date_cols = []
    header_row = None
    for r in rows[:20]:
        cols = [j for j, c in enumerate(r)
                if c is not None and str(c).strip().upper() == 'DATE']
        if cols:
            date_cols, header_row = sorted(cols), r
            break
    if not date_cols:
        return []
    roles = _hdr_role_cols(header_row, date_cols)

    headers = [i for i, r in enumerate(rows)
               if any(c is not None and _TEAM_HDR_RE.match(str(c).strip())
                      for c in r)]
    headers.append(len(rows))

    games = []
    for hi in range(len(headers) - 1):
        start, end = headers[hi], headers[hi + 1]
        name = ''
        for c in rows[start]:
            if c is not None and _TEAM_HDR_RE.match(str(c).strip()):
                name = _HDR_SUFFIX_RE.sub('', str(c).strip()).strip()
                break
            if c is not None and c.lower() in team_index.keys():
                name = team_index[c.lower()][0]
                break

        full_teams = resolve_teams(name, team_index)
        if not full_teams:
            continue

        current_sections = {}
        for row in rows[start + 1:end]:
            for dc in date_cols:
                if dc >= len(row) or row[dc] is None:
                    continue
                cell = str(row[dc]).strip()
                if cell.lower().replace(' ', '') in _SECTION_LABELS:
                    current_sections[dc] = cell
                    continue
                iso = parse_schedule_date(row[dc])
                if not iso:
                    continue
                jcol, ncol, ocol = roles[dc]
                jersey = str(row[jcol]).strip() if jcol < len(row) and row[jcol] else ''
                opponent = str(row[ocol]).strip() if ocol < len(row) and row[ocol] else ''
                comment = str(row[ncol]).strip() if ncol < len(row) and row[ncol] else ''
                if not opponent:
                    continue
                for full_team in full_teams:
                    games.append({
                        'team': full_team,
                        'date': iso,
                        'game_type': current_sections[dc] if dc in current_sections else '',
                        'opponent': opponent[:100],
                        'jersey': jersey[:100],
                        'home_game': _jersey_is_home(jersey),
                        'comment': comment[:255] if comment else '',
                    })
    return games


def parse_pergame_schedule_2009(rows, tag_teams):
    """Parse the per-game schedule sheet into a flat list of game dicts:
        {'team', 'date', 'opponent', 'jersey', 'home_game', 'comment'}
    Each per-team block starts with a team-header row carrying the season suffix.
    Drops rows that look like set-end markers or section labels."""
    team_index = build_team_index(tag_teams)

    date_cols = []
    header_row = None
    for r in rows[:20]:
        cols = [j for j, c in enumerate(r)
                if c is not None and str(c).strip().upper() == 'DATE']
        if cols:
            date_cols, header_row = sorted(cols), r
            break
    if not date_cols:
        return []
    roles = _hdr_role_cols(header_row, date_cols)

    headers = [i for i, r in enumerate(rows)
               if any(c is not None and _TEAM_HDR_RE.match(str(c).strip())
                      for c in r)]
    headers.append(len(rows))

    games = []
    for hi in range(len(headers) - 1):
        start, end = headers[hi], headers[hi + 1]
        name = ''
        for c in rows[start]:
            if c is not None and _TEAM_HDR_RE.match(str(c).strip()):
                name = _HDR_SUFFIX_RE.sub('', str(c).strip()).strip()
                break
            if c is not None and c.lower() in team_index.keys():
                name = team_index[c.lower()][0]
                break
        full_teams = resolve_teams(name, team_index)
        if not full_teams:
            continue

        current_sections = {}
        for row in rows[start + 1:end]:
            for dc in date_cols:
                if dc >= len(row) or row[dc] is None or str(row[dc]).strip().upper() == 'DATE':
                    continue
                cell = str(row[dc]).replace('2009-10', '').replace('2009', '').strip()
                if cell.lower().replace(' ', '') in _SECTION_LABELS:
                    current_sections[dc] = cell
                    continue
                cell = row[dc]
                if (re.match(_WD_RE, str(cell), re.IGNORECASE)):
                    if (re.match(_YEAR_END_MONTH_RE, str(cell), re.IGNORECASE)):
                        cell += ' 2009'
                    else:
                        cell += ' 2010'
                iso = parse_schedule_date(cell)
                if not iso:
                    continue
                jcol, ncol, ocol = roles[dc]
                jersey = str(row[jcol]).strip() if jcol < len(row) and row[jcol] else ''
                opponent = str(row[ocol]).strip() if ocol < len(row) and row[ocol] else ''
                comment = str(row[ncol]).strip() if ncol < len(row) and row[ncol] else ''
                if not opponent:
                    continue
                for full_team in full_teams:
                    games.append({
                        'team': full_team,
                        'date': iso,
                        'game_type': current_sections[dc] if dc in current_sections else '',
                        'opponent': opponent[:100],
                        'jersey': jersey[:100],
                        'home_game': _jersey_is_home(jersey),
                        'comment': comment[:255] if comment else '',
                    })
    return games

# ---------------------------------------------------------------------------
# Set-dates table reader (legacy format: one row per set, with a date range)
# ---------------------------------------------------------------------------

_TABLE_GAMES_RE = re.compile(r'\((\d+)|(\d+)\s*games?\b', re.IGNORECASE)
_SET_LABEL_RE = re.compile(
    r'^(.+?)\s+set\s*(\d+)\s*(home|away)?\s*[-:]?\s*$', re.IGNORECASE)
_SIMPLE_SET_LABEL_RE = re.compile(
    r'^set\s*(\d+)\s+(home|away|third)\s*[-:]?\s*$', re.IGNORECASE)


def _parse_set_label(text):
    """A summary-table label -> cleaned label string, or None.
    Handles 'Set 1 Home', 'Third Set 2', 'Blue Set 1', 'Set 3 Home -'."""
    t = re.sub(r'\s*[-:]\s*$', '', str(text).strip()).strip()
    if not t or re.match(r'end\b', t, re.IGNORECASE):
        return None
    if _SIMPLE_SET_LABEL_RE.match(t) or _SET_LABEL_RE.match(t):
        return t
    return None


def parse_set_dates_table(rows, season, tag_teams, set_dates_prefix):
    """Parse a 'Set Dates' tab whose per-team block ends with a summary table
    giving a date range (and game count) per set. Returns a flat list of dicts:
        {'team', 'set_label', 'game_count', 'dates'}
    'dates' preserves the original cell text (date range or single date)."""
    team_index = build_team_index(tag_teams)

    headers = [i for i, r in enumerate(rows)
               if any(c is not None and _TEAM_HDR_RE.match(str(c).strip())
                      for c in r)]
    headers.append(len(rows))

    out = []
    for hi in range(len(headers) - 1):
        start, end = headers[hi], headers[hi + 1]
        name = ''
        for c in rows[start]:
            if c is not None and _TEAM_HDR_RE.match(str(c).strip()):
                name = _HDR_SUFFIX_RE.sub('', str(c).strip()).strip()
                break
            if c is not None and c.lower() in team_index.keys():
                name = team_index[c.lower()][0]
                break

        full_teams = resolve_teams(name, team_index)
        if not full_teams:
            continue
        summary_table_start_col = None
        for i, r in enumerate(rows[start:end]):
            for j, cell in enumerate(r):
                if not cell:
                    continue
                if summary_table_start_col and j == summary_table_start_col:
                    data = []
                    for c in r[j:]:
                        if c:
                            data.append(c)
                    set_label = data[0] if len(data) > 0 else ''
                    dates = data[1] if len(data) > 1 else ''
                    if isinstance(dates, (datetime, date)):
                        dates = (dates.date() if isinstance(dates, datetime) else dates).isoformat()
                    game_count =  data[2] if len(data) > 2 else ''

                    for full_team in full_teams:
                        out.append({
                            'team': full_team,
                            'set_label': set_label[:50],
                            'game_count': game_count,
                            'dates': dates[:255],
                        })
                else:
                    if re.match(rf'^{set_dates_prefix}', str(cell).strip(), re.IGNORECASE):
                        summary_table_start_col = j
    return out
