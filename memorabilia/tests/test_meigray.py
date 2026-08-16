import hashlib
import hmac
import tempfile

from django.test import TestCase, override_settings
from django.contrib.auth.models import User
from django.urls import reverse
from django.core import mail
from django.core.files.uploadedfile import SimpleUploadedFile
from django.templatetags.static import static

from ..models import Collection, PlayerItem, PlayerGear, HockeyJersey, GeneralItem, League, GameType, UsageType, GearType, SeasonSet, UserProfile, PlayerItemImage, PlayerGearImage, GeneralItemImage, PhotoMatch, AuthSource, WantListProfile, WantList, WantListItem, WantListItemImage, OwnerInquiry, InquiryMessage, GeneralItemAuthentication
from ..relay import ingest_inbound, extract_token, strip_quoted_reply, relay_message, relay_address
from ..templatetags.memorabilia_extras import get_user_avatar_url, getmediaurl, collage_rows

from .base import BaseTestCase

import datetime as _dt
import io as _io
import json as _json
import os as _os_mg
import tempfile as _tempfile_mg

from django.core.files.base import ContentFile
from django.core.management import call_command
from django.core.management.base import CommandError


def _build_minimal_xlsx_bytes(
    tag_sheet='POP REPORT BY TAG NUMBER',
    schedule_sheet='SET DATES - JERSEY SCHEDULES',
    omit_sheets=('POP REPORT BY PLAYER',),
    tag_rows=None,
):
    """Build a tiny but valid MeiGray-ish workbook in memory.

    The tag sheet uses the canonical TAG # / Team / Player / JSY # / Color / Set
    / Size layout. The schedule sheet is intentionally empty (no team blocks),
    so parse_pergame_schedule returns no games — the test exercises the import
    plumbing without requiring a fully realistic schedule layout.
    """
    import openpyxl
    wb = openpyxl.Workbook()
    # Replace the default first sheet with the tag sheet.
    default = wb.active
    wb.remove(default)

    tag_ws = wb.create_sheet(tag_sheet)
    tag_ws.append(['TAG #', 'Team', 'Player', 'JSY #', 'Color', 'Set', 'Size', 'Comments'])
    if tag_rows is None:
        tag_rows = [
            ['W00001', 'Boston Bruins', 'Doe, John', '17', 'Black', '1', '56', 'note A'],
            ['W00002', 'Boston Bruins', 'Roe, Jane', '18', 'White', '1', '54', ''],
        ]
    for row in tag_rows:
        tag_ws.append(row)

    sched_ws = wb.create_sheet(schedule_sheet)
    # Provide a header row with at least one DATE column so parse_pergame_schedule
    # doesn't trip; no team-header rows below means it produces zero games.
    sched_ws.append(['', 'DATE', 'JSY', 'NOTES', 'Opponent'])

    for name in omit_sheets:
        wb.create_sheet(name).append(['ignored'])

    buf = _io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class MeiGrayModelTests(TestCase):
    """Basic invariants of the MeiGray models (PK, natural_key, properties)."""

    def setUp(self):
        from memorabilia.models import MeiGrayPopulationReport
        self.report = MeiGrayPopulationReport.objects.create(
            season='2023-2024', league='NHL',
        )
        # Give the file a name so __str__ etc. don't blow up.
        self.report.file.save('placeholder.xlsx', ContentFile(b'x'), save=True)

    def tearDown(self):
        # FieldFile cleanup so tests don't leave files lying around.
        if self.report.file:
            self.report.file.delete(save=False)

    def test_population_report_str(self):
        self.assertEqual(str(self.report), 'MeiGray Population Report 2023-2024 NHL')

    def test_population_report_natural_key_round_trip(self):
        from memorabilia.models import MeiGrayPopulationReport
        nk = self.report.natural_key()
        self.assertEqual(nk, ('2023-2024', 'NHL'))
        self.assertEqual(MeiGrayPopulationReport.objects.get_by_natural_key(*nk), self.report)

    def test_population_report_unique_together_season_league(self):
        from django.db import IntegrityError
        from memorabilia.models import MeiGrayPopulationReport
        with self.assertRaises(IntegrityError):
            MeiGrayPopulationReport.objects.create(season='2023-2024', league='NHL')

    def test_schedule_entry_str(self):
        from memorabilia.models import MeiGrayScheduleEntry
        sched = MeiGrayScheduleEntry.objects.create(
            season='2023-2024', league='NHL', team='Boston Bruins', report=self.report,
        )
        self.assertEqual(str(sched), 'Boston Bruins NHL 2023-2024')

    def test_schedule_game_entry_str_and_game_type_short(self):
        from memorabilia.models import MeiGrayScheduleEntry, MeiGrayScheduleGameEntry
        sched = MeiGrayScheduleEntry.objects.create(
            season='2023-2024', league='NHL', team='Boston Bruins', report=self.report,
        )
        game = MeiGrayScheduleGameEntry.objects.create(
            schedule=sched,
            opponent='Detroit',
            game_date=_dt.date(2024, 1, 5),
            jersey='Black',
            home_game=True,
            game_type='Regular Season',
        )
        self.assertIn('Boston Bruins', str(game))
        self.assertIn('Detroit', str(game))
        self.assertEqual(game.game_type_short, 'REG')

    def test_schedule_game_entry_game_type_short_variants(self):
        from memorabilia.models import MeiGrayScheduleEntry, MeiGrayScheduleGameEntry
        sched = MeiGrayScheduleEntry.objects.create(
            season='2023-2024', league='NHL', team='Boston Bruins', report=self.report,
        )
        cases = {
            'Preseason': 'PRE',
            'Regular Season': 'REG',
            'Playoffs': 'P/O',
            'Exhibition': 'EXH',
            'Round-Robin': 'RR',
            None: '',
            '': '',
            'Something Weird': 'Something Weird',
        }
        for raw, expected in cases.items():
            game = MeiGrayScheduleGameEntry(
                schedule=sched, opponent='X', game_date=_dt.date(2024, 1, 1),
                jersey='', game_type=raw,
            )
            self.assertEqual(game.game_type_short, expected, f'raw={raw!r}')

    def test_set_entry_str(self):
        from memorabilia.models import MeiGrayScheduleEntry, MeiGrayScheduleSetEntry
        sched = MeiGrayScheduleEntry.objects.create(
            season='2023-2024', league='NHL', team='Boston Bruins', report=self.report,
        )
        se = MeiGrayScheduleSetEntry.objects.create(
            schedule=sched, set_label='Set 1 Home', game_count='40', dates='Oct - Jan',
        )
        self.assertIn('Boston Bruins', str(se))
        self.assertIn('Set 1 Home', str(se))

    def test_tag_entry_str(self):
        from memorabilia.models import MeiGrayTagEntry
        tag = MeiGrayTagEntry.objects.create(
            tag_number='W00001', season='2023-24', league='NHL',
            team='Boston Bruins', player='Doe, John', color='Black',
            report=self.report,
        )
        s = str(tag)
        self.assertIn('W00001', s)
        self.assertIn('Doe, John', s)
        self.assertIn('Boston Bruins', s)

class MeiGrayRegistryTests(TestCase):
    """The season -> YearSpec registry and the check_sheets sheet auditor."""

    def test_resolve_exact_match(self):
        from memorabilia.meigray.registry import resolve
        spec = resolve('NHL', '2023-24')
        self.assertEqual(spec.manifest.tag, 'POP REPORT BY TAG NUMBER')

    def test_resolve_falls_back_to_earlier_season(self):
        from memorabilia.meigray.registry import resolve
        # 2022-23 isn't registered; should inherit from 2021-22.
        spec = resolve('NHL', '2022-23')
        self.assertEqual(spec.manifest.tag, 'POP REPORT BY TAG NUMBER')

    def test_resolve_unknown_league_raises(self):
        from memorabilia.meigray.registry import resolve
        with self.assertRaises(ValueError):
            resolve('XFL', '2024-25')

    def test_resolve_season_before_any_registered_raises(self):
        from memorabilia.meigray.registry import resolve
        with self.assertRaises(ValueError):
            resolve('NHL', '1999-00')

    def test_resolve_is_case_insensitive_on_league(self):
        from memorabilia.meigray.registry import resolve
        spec_upper = resolve('NHL', '2023-24')
        spec_lower = resolve('nhl', '2023-24')
        self.assertIs(spec_upper, spec_lower)

    def test_check_sheets_unknown_sheet_raises(self):
        from memorabilia.meigray import toolbox
        from memorabilia.meigray.registry import SheetManifest, check_sheets
        import openpyxl
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        wb.create_sheet('POP REPORT BY TAG NUMBER')
        wb.create_sheet('SOMETHING UNEXPECTED')
        manifest = SheetManifest(tag='POP REPORT BY TAG NUMBER', schedule=(), omit=())
        with self.assertRaises(ValueError) as ctx:
            check_sheets(wb, manifest, '2023-24', 'NHL')
        self.assertIn('unrecognized', str(ctx.exception).lower())

    def test_check_sheets_missing_sheet_raises(self):
        from memorabilia.meigray.registry import SheetManifest, check_sheets
        import openpyxl
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        wb.create_sheet('POP REPORT BY TAG NUMBER')
        manifest = SheetManifest(
            tag='POP REPORT BY TAG NUMBER',
            schedule=('NOT A REAL SHEET',),
            omit=(),
        )
        with self.assertRaises(ValueError) as ctx:
            check_sheets(wb, manifest, '2023-24', 'NHL')
        self.assertIn('not in the file', str(ctx.exception))

    def test_check_sheets_returns_normalized_lookup(self):
        from memorabilia.meigray.registry import SheetManifest, check_sheets, sheet
        import openpyxl
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        # Workbook sheet name has trailing whitespace / mixed case; manifest is canonical.
        wb.create_sheet('pop report by tag number ')
        manifest = SheetManifest(tag='POP REPORT BY TAG NUMBER', schedule=(), omit=())
        actual = check_sheets(wb, manifest, '2023-24', 'NHL')
        self.assertEqual(sheet(actual, 'POP REPORT BY TAG NUMBER'), 'pop report by tag number ')

    def test_every_parser_module_registered_an_entry(self):
        from memorabilia.meigray.registry import REPORTS
        # 20 parser files register one NHL season each: 2002-03 through 2021-22
        # contiguously (minus 2004-05 which was a lockout year, so 19 years
        # there), plus 2023-24 (2022-23 is unregistered and inherits 2021-22).
        nhl_seasons = {s for (lg, s) in REPORTS if lg == 'NHL'}
        self.assertEqual(len(nhl_seasons), 20)
        # Spot-check a couple to make sure the parser modules actually loaded.
        self.assertIn('2002-03', nhl_seasons)
        self.assertIn('2023-24', nhl_seasons)
        # 2004-05 was a lockout year; not registered.
        self.assertNotIn('2004-05', nhl_seasons)

class MeiGrayToolboxTests(TestCase):
    """Pure-function helpers in memorabilia.meigray.toolbox."""

    def test_short_season_truncates_full_year(self):
        from memorabilia.meigray.toolbox import short_season
        self.assertEqual(short_season('2023-2024'), '2023-24')

    def test_short_season_passthrough_when_already_short(self):
        from memorabilia.meigray.toolbox import short_season
        self.assertEqual(short_season('2023-24'), '2023-24')

    def test_norm_sheet_collapses_whitespace_and_upcases(self):
        from memorabilia.meigray.toolbox import norm_sheet
        self.assertEqual(norm_sheet('  Pop  Report\tBy Tag '), 'POP REPORT BY TAG')

    def test_detect_col_offset_finds_tag_header(self):
        from memorabilia.meigray.toolbox import detect_col_offset
        rows = [
            ('intro', None, None),
            (None, 'TAG #', 'Team'),
        ]
        self.assertEqual(detect_col_offset(rows), 1)

    def test_detect_col_offset_defaults_to_zero(self):
        from memorabilia.meigray.toolbox import detect_col_offset
        self.assertEqual(detect_col_offset([('whatever', None)]), 0)

    def test_format_set_number_handles_date_misread(self):
        from memorabilia.meigray.toolbox import format_set_number
        # Excel coerces '3-Jan' into a date; format_set_number reverses that.
        self.assertEqual(format_set_number(_dt.date(2024, 1, 3)), '3-Jan')

    def test_format_set_number_string_strip(self):
        from memorabilia.meigray.toolbox import format_set_number
        self.assertEqual(format_set_number('  1  '), '1')
        self.assertEqual(format_set_number(None), '')

    def test_parse_schedule_date_isoformats(self):
        from memorabilia.meigray.toolbox import parse_schedule_date
        self.assertEqual(parse_schedule_date(_dt.date(2024, 1, 5)), '2024-01-05')
        self.assertEqual(
            parse_schedule_date(_dt.datetime(2024, 1, 5, 19, 0)), '2024-01-05'
        )

    def test_parse_schedule_date_weekday_trailing(self):
        from memorabilia.meigray.toolbox import parse_schedule_date
        self.assertEqual(parse_schedule_date('Oct 5 2023, Wed'), '2023-10-05')

    def test_parse_schedule_date_weekday_leading(self):
        from memorabilia.meigray.toolbox import parse_schedule_date
        self.assertEqual(parse_schedule_date('Sat Jan 19, 2013'), '2013-01-19')

    def test_parse_schedule_date_returns_none_on_garbage(self):
        from memorabilia.meigray.toolbox import parse_schedule_date
        self.assertIsNone(parse_schedule_date(None))
        self.assertIsNone(parse_schedule_date(''))
        self.assertIsNone(parse_schedule_date('DATE'))
        self.assertIsNone(parse_schedule_date('not a date at all'))

    def test_build_team_index_maps_variants(self):
        from memorabilia.meigray.toolbox import build_team_index
        idx = build_team_index({'Boston Bruins', 'New York Rangers'})
        self.assertIn('boston bruins', idx)
        self.assertIn('boston', idx)
        self.assertIn('ny rangers', idx)
        self.assertIn('new york rangers', idx)

    def test_resolve_teams_strips_home_away_suffix(self):
        from memorabilia.meigray.toolbox import build_team_index, resolve_teams
        idx = build_team_index({'Boston Bruins'})
        self.assertEqual(resolve_teams('Boston Bruins - Home', idx), ['Boston Bruins'])
        self.assertEqual(resolve_teams('Boston Bruins - Away', idx), ['Boston Bruins'])
        self.assertEqual(resolve_teams('Boston Bruins - Road Whites', idx), ['Boston Bruins'])

    def test_resolve_teams_returns_empty_for_unknown(self):
        from memorabilia.meigray.toolbox import build_team_index, resolve_teams
        idx = build_team_index({'Boston Bruins'})
        self.assertEqual(resolve_teams('Mystery Team', idx), [])

    def test_jersey_is_home_logic(self):
        from memorabilia.meigray.toolbox import _jersey_is_home
        self.assertTrue(_jersey_is_home('Black'))
        self.assertTrue(_jersey_is_home(''))
        self.assertFalse(_jersey_is_home('White'))
        self.assertFalse(_jersey_is_home('Away'))
        self.assertFalse(_jersey_is_home('Road White'))

    def test_load_workbook_accepts_bytes(self):
        from memorabilia.meigray.toolbox import load_workbook
        content = _build_minimal_xlsx_bytes()
        wb = load_workbook(content)
        self.assertIn('POP REPORT BY TAG NUMBER', wb.sheetnames)

class MeiGrayTagSheetParsingTests(TestCase):
    """parse_tag_sheet + read_tag_rows against a hand-built workbook."""

    def setUp(self):
        from memorabilia.models import MeiGrayPopulationReport
        self.report = MeiGrayPopulationReport.objects.create(
            season='2023-24', league='NHL',
        )
        self.report.file.save('p.xlsx', ContentFile(b'x'), save=True)

    def tearDown(self):
        if self.report.file:
            self.report.file.delete(save=False)

    def test_read_tag_rows_parses_basic_rows(self):
        from memorabilia.meigray import toolbox
        from memorabilia.meigray.registry import SheetManifest, check_sheets

        content = _build_minimal_xlsx_bytes()
        wb = toolbox.load_workbook(content)
        manifest = SheetManifest(
            tag='POP REPORT BY TAG NUMBER',
            schedule=('SET DATES - JERSEY SCHEDULES',),
            omit=('POP REPORT BY PLAYER',),
        )
        actual = check_sheets(wb, manifest, '2023-24', 'NHL')
        entries = toolbox.read_tag_rows(wb, actual, manifest, self.report)
        self.assertEqual(len(entries), 2)
        tags = sorted(e.tag_number for e in entries)
        self.assertEqual(tags, ['W00001', 'W00002'])
        first = next(e for e in entries if e.tag_number == 'W00001')
        self.assertEqual(first.team, 'Boston Bruins')
        self.assertEqual(first.player, 'Doe, John')
        self.assertEqual(first.color, 'Black')
        self.assertEqual(first.notes, 'note A')

    def test_read_tag_rows_skips_duplicates_within_a_sheet(self):
        from memorabilia.meigray import toolbox
        from memorabilia.meigray.registry import SheetManifest, check_sheets

        content = _build_minimal_xlsx_bytes(tag_rows=[
            ['W00001', 'Boston Bruins', 'Doe, John', '17', 'Black', '1', '56', ''],
            ['W00001', 'Boston Bruins', 'Doe, John', '17', 'Black', '1', '56', 'dup'],
        ])
        wb = toolbox.load_workbook(content)
        manifest = SheetManifest(
            tag='POP REPORT BY TAG NUMBER',
            schedule=('SET DATES - JERSEY SCHEDULES',),
            omit=('POP REPORT BY PLAYER',),
        )
        actual = check_sheets(wb, manifest, '2023-24', 'NHL')
        entries = toolbox.read_tag_rows(wb, actual, manifest, self.report)
        self.assertEqual(len(entries), 1)

class MeiGrayImportCommandTests(TestCase):
    """End-to-end import via the management command using a small in-memory xlsx."""

    def setUp(self):
        self.tmpdir = _tempfile_mg.mkdtemp()

    def tearDown(self):
        import shutil
        shutil.rmtree(self.tmpdir, ignore_errors=True)
        # Delete any rows we left behind, plus their files.
        from memorabilia.models import MeiGrayPopulationReport
        for r in MeiGrayPopulationReport.objects.all():
            if r.file:
                r.file.delete(save=False)
            r.delete()

    # All the import tests use 2015-16 because its YearSpec manifest happens
    # to match the minimal workbook we build in _build_minimal_xlsx_bytes
    # (tag='POP REPORT BY TAG NUMBER', schedule='SET DATES - JERSEY SCHEDULES',
    # omit='POP REPORT BY PLAYER').
    SEASON = '2015-16'

    def _write_xlsx(self, name='test.xlsx', **kwargs):
        path = _os_mg.path.join(self.tmpdir, name)
        with open(path, 'wb') as f:
            f.write(_build_minimal_xlsx_bytes(**kwargs))
        return path

    def test_import_creates_report_and_tag_entries(self):
        from memorabilia.models import (
            MeiGrayPopulationReport, MeiGrayTagEntry, MeiGrayScheduleEntry,
        )
        path = self._write_xlsx()
        call_command('import_population_report', path, season=self.SEASON, league='NHL')

        report = MeiGrayPopulationReport.objects.get(season=self.SEASON, league='NHL')
        self.assertTrue(report.file.name)
        self.assertEqual(MeiGrayTagEntry.objects.filter(report=report).count(), 2)
        # Two tags, both Boston Bruins, so exactly one schedule row should exist.
        self.assertEqual(MeiGrayScheduleEntry.objects.filter(report=report).count(), 1)
        # Tag entry should be wired to the team's schedule row.
        tag = MeiGrayTagEntry.objects.get(tag_number='W00001')
        self.assertIsNotNone(tag.schedule)
        self.assertEqual(tag.schedule.team, 'Boston Bruins')

    def test_import_missing_file_raises_command_error(self):
        with self.assertRaises(CommandError):
            call_command(
                'import_population_report',
                _os_mg.path.join(self.tmpdir, 'does-not-exist.xlsx'),
                season=self.SEASON, league='NHL',
            )

    def test_import_unknown_sheet_rolls_back_new_report(self):
        from memorabilia.models import MeiGrayPopulationReport
        # Build a workbook that has the manifest sheets PLUS a stray sheet
        # that the registry can't account for.
        import openpyxl
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        ws = wb.create_sheet('POP REPORT BY TAG NUMBER')
        ws.append(['TAG #', 'Team', 'Player'])
        wb.create_sheet('SET DATES - JERSEY SCHEDULES').append(['DATE'])
        wb.create_sheet('POP REPORT BY PLAYER')
        wb.create_sheet('TOTALLY UNEXPECTED TAB')
        path = _os_mg.path.join(self.tmpdir, 'bad.xlsx')
        wb.save(path)

        with self.assertRaises(CommandError) as ctx:
            call_command(
                'import_population_report', path,
                season=self.SEASON, league='NHL',
            )
        # Roll-back: no report row should remain for this season+league.
        self.assertEqual(
            MeiGrayPopulationReport.objects.filter(season=self.SEASON, league='NHL').count(),
            0,
        )
        self.assertIn('unrecognized', str(ctx.exception).lower())

    def test_import_duplicate_tag_across_reports_raises(self):
        from memorabilia.models import MeiGrayPopulationReport, MeiGrayTagEntry
        # Pre-create a tag under a *different* report so the import's
        # cross-report dedup guard fires. The other report just needs to exist
        # in the DB; it's never re-parsed.
        other = MeiGrayPopulationReport.objects.create(season='1999-00', league='NHL')
        other.file.save('o.xlsx', ContentFile(b'x'), save=True)
        MeiGrayTagEntry.objects.create(
            tag_number='W00001', season='1999-00', league='NHL',
            team='Boston Bruins', player='Old, Player', color='Black',
            report=other,
        )
        path = self._write_xlsx()
        with self.assertRaises(CommandError) as ctx:
            call_command(
                'import_population_report', path,
                season=self.SEASON, league='NHL',
            )
        self.assertIn('already imported', str(ctx.exception))

    def test_import_lowercases_then_uppercases_league(self):
        from memorabilia.models import MeiGrayPopulationReport
        path = self._write_xlsx()
        call_command('import_population_report', path, season=self.SEASON, league='nhl')
        # league is stored upper-cased.
        self.assertTrue(
            MeiGrayPopulationReport.objects.filter(season=self.SEASON, league='NHL').exists()
        )

    def test_reimport_replaces_tag_rows_and_keeps_one_report(self):
        from memorabilia.models import MeiGrayPopulationReport, MeiGrayTagEntry
        path1 = self._write_xlsx(name='first.xlsx')
        call_command('import_population_report', path1, season=self.SEASON, league='NHL')
        first_count = MeiGrayPopulationReport.objects.filter(season=self.SEASON).count()
        self.assertEqual(first_count, 1)

        # Re-import with a completely different set of tag numbers.
        path2 = self._write_xlsx(name='second.xlsx', tag_rows=[
            ['Z00099', 'Boston Bruins', 'Sub, Player', '99', 'Black', '1', '54', ''],
        ])
        call_command('import_population_report', path2, season=self.SEASON, league='NHL')

        # Still exactly one report row for this season+league.
        self.assertEqual(
            MeiGrayPopulationReport.objects.filter(season=self.SEASON, league='NHL').count(),
            1,
        )
        # Old tag numbers are gone; new one is present.
        tags = set(MeiGrayTagEntry.objects.values_list('tag_number', flat=True))
        self.assertNotIn('W00001', tags)
        self.assertIn('Z00099', tags)

class MeiGrayExportCommandTests(TestCase):
    """The export_population_report_data management command."""

    def setUp(self):
        self.tmpdir = _tempfile_mg.mkdtemp()
        from memorabilia.models import (
            MeiGrayPopulationReport, MeiGrayScheduleEntry,
            MeiGrayScheduleGameEntry, MeiGrayScheduleSetEntry, MeiGrayTagEntry,
        )
        self.report = MeiGrayPopulationReport.objects.create(
            season='2023-24', league='NHL',
        )
        self.report.file.save('r.xlsx', ContentFile(b'x'), save=True)
        self.sched = MeiGrayScheduleEntry.objects.create(
            season='2023-24', league='NHL', team='Boston Bruins', report=self.report,
        )
        MeiGrayScheduleGameEntry.objects.create(
            schedule=self.sched, opponent='Detroit', game_date=_dt.date(2023, 10, 5),
            jersey='Black', home_game=True, game_type='Regular Season',
        )
        MeiGrayScheduleSetEntry.objects.create(
            schedule=self.sched, set_label='Set 1 Home', game_count='40', dates='Oct - Jan',
        )
        MeiGrayTagEntry.objects.create(
            tag_number='W00001', season='2023-24', league='NHL',
            team='Boston Bruins', player='Doe, John', color='Black',
            report=self.report, schedule=self.sched,
        )

    def tearDown(self):
        import shutil
        shutil.rmtree(self.tmpdir, ignore_errors=True)
        if self.report.file:
            self.report.file.delete(save=False)

    def test_export_writes_loaddata_compatible_fixture(self):
        out = _os_mg.path.join(self.tmpdir, 'out.json')
        call_command(
            'export_population_report_data',
            season='2023-24', league='NHL', output=out,
        )
        self.assertTrue(_os_mg.path.exists(out))
        with open(out) as f:
            data = _json.load(f)
        models = {d['model'] for d in data}
        self.assertIn('memorabilia.meigraypopulationreport', models)
        self.assertIn('memorabilia.meigrayscheduleentry', models)
        self.assertIn('memorabilia.meigrayschedulegameentry', models)
        self.assertIn('memorabilia.meigrayschedulesetentry', models)
        self.assertIn('memorabilia.meigraytagentry', models)

    def test_export_uses_natural_keys_for_report_foreignkey(self):
        out = _os_mg.path.join(self.tmpdir, 'out.json')
        call_command(
            'export_population_report_data',
            season='2023-24', league='NHL', output=out,
        )
        with open(out) as f:
            data = _json.load(f)
        sched = next(d for d in data if d['model'] == 'memorabilia.meigrayscheduleentry')
        # FK to MeiGrayPopulationReport should be serialized as its natural key.
        self.assertEqual(sched['fields']['report'], ['2023-24', 'NHL'])

    def test_export_missing_season_raises(self):
        out = _os_mg.path.join(self.tmpdir, 'out.json')
        with self.assertRaises(CommandError):
            call_command(
                'export_population_report_data',
                season='1900-01', league='NHL', output=out,
            )

    def test_export_then_reload_round_trips(self):
        from memorabilia.models import (
            MeiGrayPopulationReport, MeiGrayScheduleEntry,
            MeiGrayScheduleGameEntry, MeiGrayScheduleSetEntry, MeiGrayTagEntry,
        )
        out = _os_mg.path.join(self.tmpdir, 'out.json')
        call_command(
            'export_population_report_data',
            season='2023-24', league='NHL', output=out,
        )
        # Wipe everything and reload from the exported fixture.
        MeiGrayTagEntry.objects.all().delete()
        MeiGrayScheduleGameEntry.objects.all().delete()
        MeiGrayScheduleSetEntry.objects.all().delete()
        MeiGrayScheduleEntry.objects.all().delete()
        if self.report.file:
            self.report.file.delete(save=False)
        MeiGrayPopulationReport.objects.all().delete()

        call_command('loaddata', out)

        self.assertEqual(MeiGrayPopulationReport.objects.count(), 1)
        self.assertEqual(MeiGrayScheduleEntry.objects.count(), 1)
        self.assertEqual(MeiGrayScheduleGameEntry.objects.count(), 1)
        self.assertEqual(MeiGrayScheduleSetEntry.objects.count(), 1)
        self.assertEqual(MeiGrayTagEntry.objects.count(), 1)
        # Tag's schedule FK survives the round trip.
        tag = MeiGrayTagEntry.objects.get(tag_number='W00001')
        self.assertIsNotNone(tag.schedule)
        self.assertEqual(tag.schedule.team, 'Boston Bruins')

class MeiGrayFixtureLoadTests(TestCase):
    """Smoke-test that a real shipping fixture deserializes end-to-end."""

    def test_2023_2024_fixture_loads(self):
        from memorabilia.models import (
            MeiGrayPopulationReport, MeiGrayScheduleEntry,
            MeiGrayScheduleGameEntry, MeiGrayTagEntry,
        )
        call_command('loaddata', '2023-2024-NHL.json', verbosity=0)
        report = MeiGrayPopulationReport.objects.get(season='2023-2024', league='NHL')
        self.assertTrue(MeiGrayScheduleEntry.objects.filter(report=report).exists())
        # Spot-check counts match the fixture file.
        self.assertEqual(MeiGrayTagEntry.objects.filter(report=report).count(), 3281)
        self.assertEqual(MeiGrayScheduleGameEntry.objects.count(), 540)

class DownloadPopulationReportViewTests(BaseTestCase):
    """The download_population_report view."""

    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()
        from memorabilia.models import MeiGrayPopulationReport
        cls.report = MeiGrayPopulationReport.objects.create(
            season='2023-24', league='NHL',
        )
        cls.report.file.save(
            'sample_population_report.xlsx',
            ContentFile(_build_minimal_xlsx_bytes()),
            save=True,
        )

    @classmethod
    def tearDownClass(cls):
        if cls.report.file:
            cls.report.file.delete(save=False)
        super().tearDownClass()

    def test_anonymous_redirected_to_login(self):
        response = self.client.get(
            reverse('memorabilia:download_population_report', kwargs={'league': 'NHL', 'season': '2023-24'})
        )
        self.assertEqual(response.status_code, 302)
        self.assertIn('/accounts/login/', response.url)

    def test_authenticated_user_can_download(self):
        self.client.force_login(self.owner)
        response = self.client.get(
            reverse('memorabilia:download_population_report', kwargs={'league': 'NHL', 'season': '2023-24'})
        )
        self.assertEqual(response.status_code, 200)
        # FileResponse sets attachment + filename.
        self.assertIn('attachment', response.get('Content-Disposition', ''))
        self.assertIn('.xlsx', response.get('Content-Disposition', ''))
        # Drain the streaming response so the underlying file handle is closed
        # before tearDown deletes the file (avoids a Windows-style ResourceWarning).
        if hasattr(response, 'streaming_content'):
            for _ in response.streaming_content:
                pass

    def test_unknown_season_returns_404(self):
        self.client.force_login(self.owner)
        response = self.client.get(
            reverse('memorabilia:download_population_report', kwargs={'league': 'NHL', 'season': '1900-01'})
        )
        self.assertEqual(response.status_code, 404)

class HockeyJerseyMeiGrayDetailContextTests(BaseTestCase):
    """The hockey jersey detail view exposes matching MeiGrayTagEntry rows
    when the jersey has a MEIGRAY-issued authentication number."""

    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()
        from memorabilia.models import (
            CoaType, AuthSource, MeiGrayPopulationReport,
            MeiGrayScheduleEntry, MeiGrayTagEntry, PlayerGearAuthentication,
        )
        cls.coa_paper = CoaType.objects.get_or_create(
            key='paper', defaults={'name': 'Paper'}
        )[0]
        cls.meigray = AuthSource.objects.get_or_create(
            key='MEIGRAY', defaults={'name': 'MeiGray'}
        )[0]
        cls.other_source = AuthSource.objects.get_or_create(
            key='BAS', defaults={'name': 'BAS'}
        )[0]
        cls.report = MeiGrayPopulationReport.objects.create(
            season='2023-24', league='NHL',
        )
        cls.report.file.save('r.xlsx', ContentFile(b'x'), save=True)
        cls.sched = MeiGrayScheduleEntry.objects.create(
            season='2023-24', league='NHL', team='Boston Bruins', report=cls.report,
        )
        cls.tag = MeiGrayTagEntry.objects.create(
            tag_number='WMGZ01', season='2023-24', league='NHL',
            team='Boston Bruins', player='Doe, John', color='Black',
            report=cls.report, schedule=cls.sched,
        )
        cls.jersey_with_tag = HockeyJersey.objects.create(
            title='MeiGray Jersey', description='', collection=cls.collection,
            league='NHL', player='Doe, John', brand='Adidas', size='56', season='2023-24',
            game_type=cls.game_type, usage_type=cls.usage_type,
        )
        PlayerGearAuthentication.objects.create(
            collectible=cls.jersey_with_tag, auth_type=cls.coa_paper,
            number='WMGZ01', issuer=cls.meigray,
        )

    @classmethod
    def tearDownClass(cls):
        if cls.report.file:
            cls.report.file.delete(save=False)
        super().tearDownClass()

    def _detail_url(self, pk):
        return reverse(
            'memorabilia:collectible',
            kwargs={
                'collection_id': self.collection.id,
                'collectible_type': 'hockeyjersey',
                'pk': pk,
            },
        )

    def test_meigray_entries_present_when_auth_matches(self):
        response = self.client.get(self._detail_url(self.jersey_with_tag.id))
        self.assertEqual(response.status_code, 200)
        entries = response.context.get('meigray_entries')
        self.assertIsNotNone(entries)
        self.assertEqual([e.tag_number for e in entries], ['WMGZ01'])

    def test_meigray_entries_absent_for_jersey_without_meigray_auth(self):
        # The base fixture's HockeyJersey has no MEIGRAY authentication.
        response = self.client.get(self._detail_url(self.hockey_jersey.id))
        self.assertEqual(response.status_code, 200)
        self.assertNotIn('meigray_entries', response.context)

    def test_meigray_entries_absent_when_auth_issuer_is_not_meigray(self):
        from memorabilia.models import PlayerGearAuthentication
        jersey = HockeyJersey.objects.create(
            title='Other Auth Jersey', description='', collection=self.collection,
            league='NHL', player='X', brand='CCM', size='56', season='2023-24',
            game_type=self.game_type, usage_type=self.usage_type,
        )
        # Non-MEIGRAY issuer with a tag number that *happens* to exist in MeiGrayTagEntry.
        PlayerGearAuthentication.objects.create(
            collectible=jersey, auth_type=self.coa_paper,
            number='WMGZ01', issuer=self.other_source,
        )
        response = self.client.get(self._detail_url(jersey.id))
        self.assertEqual(response.status_code, 200)
        self.assertNotIn('meigray_entries', response.context)
